from __future__ import annotations

from pathlib import Path
import shutil

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\CasaT\Downloads\sar hopton")
OUT = Path(r"C:\Users\CasaT\quaerensclaims\document_work\caravan_claim_review\sar_hopton_spreadsheet_digest.txt")
TEMP = OUT.with_suffix(".xlsx")


def main() -> None:
    shutil.copyfile(SOURCE, TEMP)
    wb = load_workbook(TEMP, data_only=True, read_only=True)
    chunks = []
    for ws in wb.worksheets:
        chunks.append(f"\n===== SHEET: {ws.title} =====")
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if vals:
                chunks.append(" | ".join(vals))
    OUT.write_text("\n".join(chunks), encoding="utf-8", errors="ignore")
    print(f"Sheets: {', '.join(ws.title for ws in wb.worksheets)}")
    print(f"Digest: {OUT}")


if __name__ == "__main__":
    main()
